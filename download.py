from random import SystemRandom
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import openpyxl
import requests
import decimal
import base64
import tqdm
import time
import csv
import os

no_match = []
decimal.getcontext().rounding = "ROUND_HALF_UP"

dir_download = r'2023年1月流量图'
root = os.path.join(os.getcwd(), dir_download)
if not os.path.exists(root):
    os.mkdir(root)

xlfile = r'2023年1月流量图.xlsx'
xlfile = os.path.join(root, xlfile)

root = os.path.join(root, 'NODE_ROOT')
if not os.path.exists(root):
    os.mkdir(root)

# 需要出的图对应的 id
NODE_TREE = {
    "ROOT_DIR1": {
        "SUB_DIR1": [GRAPH_ID1, GRAPH_ID2, GRAPH_ID3]
    },
    "ROOT_DIR2": {
        "SUB_DIR1": [GRAPH_ID4, GRAPH_ID5, GRAPH_ID6]
}


def get_jpg_csv(url, header, graph_start, graph_end):
    png_url = url + '/graph_image.php?'
    csv_url = url + '/graph_xport.php?'

    graph_start = graph_start
    graph_end = graph_end

    if not os.path.exists(xlfile):
        wb = Workbook()

    for node, customers in NODE.items():
        print(node)
        node_dir = os.path.join(root, node)
        os.mkdir(node_dir)

        # 创建以节点为命名的工作表
        ws = wb.create_sheet(node)
        start = 1
        # print(f'A{start}      for node, customers in NODE.items():')
        # print(node_dir)
        for customer, png_list in customers.items():
            print(customer)
            customer_dir = os.path.join(node_dir, customer)
            os.mkdir(customer_dir)

            # 在插入的图片上面注明客户名称
            ws[f'A{start}'].value = customer
            start = start + 1
            # print(f'A{start}       for customer, png_list in customers.items():')
            print(customer_dir)
            for png in png_list:
                print(png)
                param = {"local_graph_id": png, "graph_start": graph_start, "graph_end": graph_end}

                while True:
                    try:
                        csv_response = requests.get(url=csv_url, headers=headers, params=param, timeout=None,
                                                    stream=True)
                        png_response = requests.get(url=png_url, headers=headers, params=param, timeout=None,
                                                    stream=True)
                        break
                    except requests.exceptions.ConnectionError:
                        time.sleep(5)
                time.sleep(23)
                # 从response header 中获取csv文件名
                csv_name = csv_response.headers.get("Content-Disposition").split("=")[1]
                # 替换文件名中非法字符，不然保存时会出错
                csv_name = csv_name.replace('"', "").replace("/", "-").replace("|", "+").replace("*", "x")
                # 重新编码，防止乱码
                csv_name = csv_name.encode('ISO-8859-1').decode('utf-8')
                csv_name = csv_name.replace('"', "").replace("'", "")
                csv_name = f'{png}-{csv_name}'

                # png文件名前缀和csv文件名相同
                png_name = csv_name.replace('"', "").replace("'", "").replace("csv", "png")

                # 拼接保存路径
                csv_file = os.path.join(customer_dir, csv_name)
                png_file = os.path.join(customer_dir, png_name)

                with open(csv_file, 'wb+') as csv_save, open(png_file, 'wb+') as png_save:
                    csv_save.write(csv_response.content)
                    png_save.write(png_response.content)

                if not os.path.exists(png_file):
                    print(f'{png} 下载失败！')

                image = Image(png_file)
                image.width = int(image.width * 1)
                image.height = int(image.height * 1)

                width = int(image.width / 37.8)
                height = int(image.height / 37.8)

                row = int(height / 0.51) + 1

                ws.add_image(image, f'A{start}')

                start = (row + 1) + start
                start += 1
                # print(f'A{start}      ws.add_image')

                os.chdir(os.path.dirname(csv_file))
                nth_percent(csv_file)
            start += 2
    wb.save(xlfile)


# def insert_png(png_file,):
#     file = 'XX年XX月流量图.xlsx'
#     if not os.path.exists(file):
#         wb = Workbook()
#         ws = wb.create_sheet(f'{node}')
#
#         # 要插入表格的图片
#         image = Image(png_file)
#         # 插入表格的图片缩小0.9 倍
#         image.width = int(image.width * 0.9)
#         image.height = int(image.height * 0.9)


# 生成cookie
def gen_salt(length):
    """
    :param length: 要生成cookie 的长度
    """
    salt_chars = "abcdefghijklmnopqrstuvwxyz0123456789"
    sys_random = SystemRandom()
    return "".join(sys_random.choice(salt_chars) for i in range(length))


# basic_auth 编码
def gen_auth(basic_user, basic_pass):
    return str(base64.b64encode(f'{basic_user}:{basic_pass}'.encode('utf-8')), 'utf-8')


# 登录 cactiez
def cacti_login(
        host,
        basic_user,
        basic_pass,
        cacti_user,
        cacti_pass,
        port=80
):
    """
    :param host: cactiez 的IP地址
    :param port: cactiez 的端口，默认80 端口
    :param basic_user: basic auth 的用户
    :param basic_pass: basic auth 用户的密码
    :param cacti_user: cactiez 的用户
    :param cacti_pass: cactiez 的密码
    :return : header
    """

    cookie = gen_salt(26)
    auth = gen_auth(basic_user, basic_pass)

    data = {
        'action': 'login',
        'login_username': cacti_user,
        'login_password': cacti_pass
    }

    headers = {"Cookie": f'Cacti={cookie}', "Authorization": f'Basic {auth}',
               "User-Agent": "Mozilla/5.0 (Windows NT 10.0; "
                             "Win64; x64; rv:105.0) "
                             "Gecko/20100101 Firefox/105.0"}
    if not port == 80:
        host = f'http://{host}:{port}'
    login = f'{host}/index.php'

    # response = requests.post(url=login, auth=HTTPBasicAuth(basic_user, basic_pass), data=data)

    response = requests.post(url=login, headers=headers, data=data)
    if not response.ok:
        print(response.status_code)
        return

    # print(response.status_code)
    # print(response.headers)
    # cookies = requests.utils.dict_from_cookiejar(response.cookies)
    # print(cookies)

    # logout = requests.get(url=logout, headers=headers)
    # print(logout.status_code)

    return host, headers


# 退出登录
def cacti_logout(host, headers):
    url = f'{host}/logout.php'
    logout = requests.get(url=url, headers=headers)
    if not logout.cookies == headers['Cookie']:
        print(f'CactiEZ 退出成功')


# 验证原始数据
def nth_percent(csv_file):
    csv_file = csv_file
    csv_head = []
    csv_tail = []

    left_list = []
    left_total = 0
    right_list = []
    right_total = 0

    invaild = 0

    title_file = os.path.basename(csv_file).split('.csv')[0] + '.txt'
    # print(title_file)
    csv_dir = os.path.dirname(csv_file)
    exists = os.path.exists(csv_file)
    if exists:
        title_open = open(title_file, mode='wt', encoding='utf-8')
        csv_open = open(csv_file, mode='rt', encoding='utf-8')

        for i, line in enumerate(csv_open):
            while i <= 9:
                # print(line.strip())
                title_open.write(line)
                break

        csv_open.close()
        title_open.close()
    else:
        print(csv_file + '\n File not found')

    with open(csv_file, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        # row = [row for row in reader]
        for i, rows in enumerate(reader):
            if i < 11:
                row = rows
                csv_head.append(row)
            if i >= 11:
                row = rows
                csv_tail.append(row)

    count_rows = len(csv_tail)
    ports_num = (len(csv_tail[0]) - 1)

    left_port = int((ports_num / 2))
    # print('总行数 {}行, 取 {}行'.format(count_rows, int(count_rows * 0.05)))
    # print('一共 {} 个端口'.format(ports_num))

    line = csv_tail[0]
    total_lines = csv_head[5][1]
    graph_nth = csv_head[8][1]
    graph_nth = decimal.Decimal(str(graph_nth)).quantize(decimal.Decimal("0.1"))

    select_line = int(float(total_lines) * 0.05)
    if left_port == 1:
        sub = True
        max_port = []
        for every_line in csv_tail:
            mid = 0
            if float(every_line[1]) > float(every_line[2]):
                mid = float(every_line[1])
            else:
                mid = float(every_line[2])
            max_port.append(mid)
        max_port.sort(reverse=True)
        for i in range(select_line - 1, select_line + 2):
            value = decimal.Decimal(str(max_port[i])).quantize(decimal.Decimal("0.1"))
            # print('{} 行：{}'.format(i, value))
            if value == graph_nth:
                sub = True
                break
            else:
                sub = False

        # print(sub)

        with open(title_file, 'a+', encoding='utf-8') as f:
            f.write('总行数 {}行, 取 {}行\n'.format(count_rows, int(count_rows * 0.05)))
            f.write('一共 {} 个端口\n'.format(ports_num))
            f.write('95值保留一位小数: {}\n'.format(graph_nth))
            for i in range(select_line - 1, select_line + 2):
                value = decimal.Decimal(str(max_port[i])).quantize(decimal.Decimal("0.1"))
                f.write('{} 行：{}\n'.format(i, value))
            f.write(str(sub))

    else:
        for every_line in csv_tail:
            left_sum = 0
            right_sum = 0
            for i in range(1, left_port + 1):
                left_sum += float(every_line[i])
            left_list.append(left_sum)
            for i in range(left_port + 1, ports_num + 1):
                right_sum += float(every_line[i])
            right_list.append(right_sum)

        for ele in range(0, len(left_list)):
            left_total = left_total + left_list[ele]
        for ele in range(0, len(right_list)):
            right_total = right_total + right_list[ele]
        # print(left_total)
        # print(right_total)
        left_list.sort(reverse=True)
        right_list.sort(reverse=True)
        if left_total > right_total:
            for i in range(select_line - 1, select_line + 2):
                value = decimal.Decimal(str(left_list[i])).quantize(decimal.Decimal("0.1"))
                # print('{} 行：{}'.format(i, value))
                if value == graph_nth:
                    sub = True
                    break
                else:
                    sub = False
            # print(sub)
            with open(title_file, 'a+', encoding='utf-8') as f:
                f.write('总行数 {}行, 取 {}行\n'.format(count_rows, int(count_rows * 0.05)))
                f.write('一共 {} 个端口\n'.format(ports_num))
                f.write('95值保留一位小数: {}\n'.format(graph_nth))
                for i in range(select_line - 1, select_line + 2):
                    value = decimal.Decimal(str(left_list[i])).quantize(decimal.Decimal("0.1"))
                    f.write('{} 行：{}\n'.format(i, value))
                f.write(str(sub))
        else:
            for i in range(select_line - 1, select_line + 2):
                value = decimal.Decimal(str(right_list[i])).quantize(decimal.Decimal("0.1"))
                # print('{} 行：{}'.format(i, value))
                if value == graph_nth:
                    sub = True
                    break
                else:
                    sub = False
                    no_match.append(title_file)

            # print(sub)

            with open(title_file, 'a+', encoding='utf-8') as f:
                f.write('总行数 {}行, 取 {}行\n'.format(count_rows, int(count_rows * 0.05)))
                f.write('一共 {} 个端口\n'.format(ports_num))
                f.write('95值保留一位小数: {}\n'.format(graph_nth))
                for i in range(select_line - 1, select_line + 2):
                    value = decimal.Decimal(str(right_list[i])).quantize(decimal.Decimal("0.1"))
                    f.write('{} 行：{}\n'.format(i, value))
                f.write(str(sub))
            # print(right_list)


if __name__ == '__main__':

    url_cookie = cacti_login(host='http://CACTI_IP', port=80, basic_user='Username', basic_pass='Password',
                             cacti_user='CACTI_USERNAME',
                             cacti_pass='CACTI_PASSWORD')

    headers = url_cookie[1]
    url = url_cookie[0]

    #
    start_time = "2023-1-1 00:00:00"
    end_time = "2023-1-31 23:59:59"
    start_time_Arry = time.strptime(start_time, '%Y-%m-%d %H:%M:%S')
    end_time_Arry = time.strptime(end_time, '%Y-%m-%d %H:%M:%S')
    start_timestamp = time.mktime(start_time_Arry)
    end_timestamp = time.mktime(end_time_Arry)

    start_time = int(start_timestamp)
    end_time = int(end_timestamp)

    get_jpg_csv(url, headers, start_time, end_time)

    # 退出登录
    cacti_logout(url, headers)
